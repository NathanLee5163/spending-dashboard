"""Core logic for organizing transaction CSV files."""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import BinaryIO, TextIO

# Common categories from credit card / budgeting exports (Monarch, Mint, etc.)
STANDARD_CATEGORIES = [
    "Auto & transport",
    "Bills & utilities",
    "Business",
    "Charity & donations",
    "Childcare & education",
    "Clothing",
    "Credit card payment",
    "Drinks & dining",
    "Entertainment",
    "Fees & charges",
    "Fitness",
    "Groceries",
    "Gifts & donations",
    "Health & wellness",
    "Home",
    "Income",
    "Insurance",
    "Investments",
    "Loan payment",
    "Other",
    "Personal care",
    "Pets",
    "Phone & internet",
    "Rent & mortgage",
    "Shopping",
    "Subscriptions",
    "Taxes",
    "Transfer",
    "Travel & vacation",
    "Cash & ATM",
    "Uncategorized",
]

CATEGORY_COLORS = {
    "Auto & transport": "#f59e0b",
    "Bills & utilities": "#6366f1",
    "Business": "#8b5cf6",
    "Charity & donations": "#ec4899",
    "Childcare & education": "#14b8a6",
    "Clothing": "#d946ef",
    "Credit card payment": "#64748b",
    "Drinks & dining": "#ef4444",
    "Entertainment": "#a855f7",
    "Fees & charges": "#78716c",
    "Fitness": "#22c55e",
    "Groceries": "#10b981",
    "Gifts & donations": "#f472b6",
    "Health & wellness": "#06b6d4",
    "Home": "#84cc16",
    "Income": "#34d399",
    "Insurance": "#0ea5e9",
    "Investments": "#2dd4bf",
    "Loan payment": "#94a3b8",
    "Other": "#9ca3af",
    "Personal care": "#fb7185",
    "Pets": "#fbbf24",
    "Phone & internet": "#60a5fa",
    "Rent & mortgage": "#4ade80",
    "Shopping": "#3b82f6",
    "Subscriptions": "#c084fc",
    "Taxes": "#f97316",
    "Transfer": "#6b7280",
    "Travel & vacation": "#38bdf8",
    "Cash & ATM": "#a3a3a3",
    "Uncategorized": "#475569",
}


def parse_amount(value: str) -> float:
    cleaned = value.strip().replace("$", "").replace(",", "")
    return round(float(cleaned), 2)


def format_money(amount: float, signed: bool = False) -> str:
    value = f"${abs(amount):,.2f}"
    if not signed:
        return value
    if amount > 0:
        return f"+{value}"
    if amount < 0:
        return f"-{value}"
    return value


DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d")
EXPENSE_TYPES = {"expense"}

DATE_ALIASES = (
    "transaction date",
    "trans date",
    "posted date",
    "post date",
    "posting date",
    "itransaction date",
    "date",
)
DESCRIPTION_ALIASES = (
    "description",
    "payee",
    "transaction description",
    "statement description",
    "details",
    "memo",
    "merchant",
    "name",
)
CATEGORY_ALIASES = ("category",)
TYPE_ALIASES = ("type", "transaction type", "status")
AMOUNT_ALIASES = ("amount",)
DEBIT_ALIASES = ("debit", "debits", "debit amount", "charge", "withdrawals", "withdrawal")
CREDIT_ALIASES = ("credit", "credits", "credit amount", "deposits", "deposit")
ACCOUNT_ALIASES = ("account", "card no.", "card no", "card number")

# Bank & budgeting-app categories mapped to a shared standard
CATEGORY_ALIASES_MAP = {
    "food & drink": "Drinks & dining",
    "food and drink": "Drinks & dining",
    "dining": "Drinks & dining",
    "restaurants": "Drinks & dining",
    "drinks & dining": "Drinks & dining",
    "gas": "Auto & transport",
    "automotive": "Auto & transport",
    "auto & transport": "Auto & transport",
    "transportation": "Auto & transport",
    "travel": "Travel & vacation",
    "travel & vacation": "Travel & vacation",
    "bills & utilities": "Bills & utilities",
    "bills and utilities": "Bills & utilities",
    "utilities": "Bills & utilities",
    "health & wellness": "Health & wellness",
    "health and wellness": "Health & wellness",
    "healthcare": "Health & wellness",
    "medical": "Health & wellness",
    "professional services": "Business",
    "business services": "Business",
    "groceries": "Groceries",
    "grocery": "Groceries",
    "shopping": "Shopping",
    "merchandise": "Shopping",
    "entertainment": "Entertainment",
    "subscriptions": "Subscriptions",
    "fees & charges": "Fees & charges",
    "fees": "Fees & charges",
    "personal care": "Personal care",
    "home": "Home",
    "insurance": "Insurance",
    "education": "Childcare & education",
    "childcare & education": "Childcare & education",
    "gifts & donations": "Gifts & donations",
    "charity": "Charity & donations",
    "pets": "Pets",
    "fitness": "Fitness",
    "income": "Income",
    "other": "Other",
    "cash & atm": "Cash & ATM",
    "atm": "Cash & ATM",
}

TYPE_ALIASES_MAP = {
    "sale": "Expense",
    "debit": "Expense",
    "purchase": "Expense",
    "charge": "Expense",
    "withdrawal": "Expense",
    "expense": "Expense",
    "fee": "Expense",
    "adjustment": "Expense",
    "payment": "Transfer",
    "transfer": "Transfer",
    "credit card payment": "Transfer",
    "cc payment": "Transfer",
    "loan payment": "Transfer",
    "return": "Return",
    "refund": "Return",
    "reversal": "Return",
    "deposit": "Income",
    "income": "Income",
    "credit": "Transfer",
}


def clean_header(name: str) -> str:
    return name.strip().lstrip("\ufeff").strip('"')


def normalize_date(value: str) -> str:
    cleaned = value.strip().strip('"')
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return cleaned


def parse_month(date_value: str) -> str:
    normalized = normalize_date(date_value)
    try:
        return datetime.strptime(normalized, "%Y-%m-%d").strftime("%Y-%m")
    except ValueError:
        return "Unknown"


def normalize_category(name: str) -> str:
    cleaned = name.strip() or "Uncategorized"
    return CATEGORY_ALIASES_MAP.get(cleaned.lower(), cleaned)


def _match_column(headers: list[str], aliases: tuple[str, ...]) -> str | None:
    normalized = {clean_header(header).lower(): header for header in headers if header}

    for alias in aliases:
        if alias in normalized:
            return normalized[alias]

    for alias in aliases:
        for key, original in normalized.items():
            if alias in key:
                return original

    return None


def _detect_columns(headers: list[str]) -> dict[str, str | None]:
    return {
        "date": _match_column(headers, DATE_ALIASES),
        "description": _match_column(headers, DESCRIPTION_ALIASES),
        "category": _match_column(headers, CATEGORY_ALIASES),
        "type": _match_column(headers, TYPE_ALIASES),
        "amount": _match_column(headers, AMOUNT_ALIASES),
        "debit": _match_column(headers, DEBIT_ALIASES),
        "credit": _match_column(headers, CREDIT_ALIASES),
        "account": _match_column(headers, ACCOUNT_ALIASES),
    }


def _find_header_index(lines: list[str]) -> int:
    for index, line in enumerate(lines[:30]):
        if not line.strip():
            continue

        try:
            row = next(csv.reader([line]))
        except csv.Error:
            continue

        lowered = [clean_header(cell).lower() for cell in row]
        has_date = any(
            any(alias == cell or alias in cell for alias in DATE_ALIASES)
            for cell in lowered
        )
        has_money = any(
            any(alias == cell or alias in cell for alias in AMOUNT_ALIASES + DEBIT_ALIASES + CREDIT_ALIASES)
            for cell in lowered
        )
        has_description = any(
            any(alias == cell or alias in cell for alias in DESCRIPTION_ALIASES)
            for cell in lowered
        )

        if has_date and (has_money or has_description):
            return index

    return 0


def _has_value(value: str | None) -> bool:
    return bool(value and value.strip())


def _parse_optional_amount(value: str | None) -> float | None:
    if not _has_value(value):
        return None
    try:
        return parse_amount(value or "")
    except ValueError:
        return None


def _resolve_amount(row: dict[str, str], columns: dict[str, str | None]) -> tuple[str, str]:
    amount_key = columns.get("amount")
    debit_key = columns.get("debit")
    credit_key = columns.get("credit")

    if amount_key and _has_value(row.get(amount_key)):
        return row[amount_key].strip(), "amount"

    debit = _parse_optional_amount(row.get(debit_key or ""))
    credit = _parse_optional_amount(row.get(credit_key or ""))

    if debit is not None and debit != 0:
        return str(-abs(debit)), "debit"
    if credit is not None and credit != 0:
        return str(abs(credit)), "credit"
    return "0", "none"


def _looks_like_refund(description: str) -> bool:
    upper = description.upper()
    return any(token in upper for token in ("REFUND", "RETURN", "REVERSAL", "CREDIT ADJ"))


def _looks_like_payment(description: str) -> bool:
    upper = description.upper()
    return any(
        token in upper
        for token in (
            "PAYMENT",
            "AUTOPAY",
            "AUTO PAY",
            "THANK YOU",
            "ONLINE PAYMENT",
            "CARD PAYMENT",
            "TRANSFER TO",
            "TRANSFER FROM",
        )
    )


def _resolve_type(
    row: dict[str, str],
    columns: dict[str, str | None],
    amount: str,
    amount_source: str,
    description: str,
) -> str:
    type_key = columns.get("type")
    if type_key and _has_value(row.get(type_key)):
        raw_type = row.get(type_key, "").strip().lower()
        mapped = TYPE_ALIASES_MAP.get(raw_type)
        if mapped:
            return mapped
        if raw_type:
            return raw_type.title()

    if _looks_like_payment(description):
        return "Transfer"

    if amount_source == "debit":
        return "Expense"

    if amount_source == "credit":
        if _looks_like_refund(description):
            return "Return"
        return "Transfer"

    try:
        numeric_amount = parse_amount(amount)
    except ValueError:
        return "Unknown"

    if numeric_amount < 0:
        return "Expense"
    if numeric_amount > 0:
        if _looks_like_refund(description):
            return "Return"
        return "Transfer"
    return "Unknown"


def _resolve_account(row: dict[str, str], columns: dict[str, str | None]) -> str:
    account_key = columns.get("account")
    if account_key and _has_value(row.get(account_key)):
        value = row[account_key].strip()
        if "card" in clean_header(account_key).lower():
            return f"Card ending {value}"
        return value
    return ""


def _read_csv_text(source: Path | str | TextIO) -> str:
    if isinstance(source, Path):
        return source.read_text(encoding="utf-8-sig")
    if isinstance(source, str):
        return source
    return source.read()


def normalize_row(row: dict[str, str], columns: dict[str, str | None]) -> dict[str, str]:
    date_key = columns.get("date")
    description_key = columns.get("description")
    category_key = columns.get("category")

    description = row.get(description_key or "", "").strip()
    amount, amount_source = _resolve_amount(row, columns)

    return {
        "Date": normalize_date(row.get(date_key or "", "")),
        "Description": description,
        "Type": _resolve_type(row, columns, amount, amount_source, description),
        "Category": normalize_category(row.get(category_key or "", "")),
        "Amount": amount,
        "Account": _resolve_account(row, columns),
    }


def parse_csv_rows(raw_text: str) -> list[dict[str, str]]:
    lines = raw_text.splitlines()
    if not lines:
        raise ValueError("No rows found in CSV file.")

    header_index = _find_header_index(lines)
    reader = csv.DictReader(lines[header_index:])
    headers = [clean_header(header) for header in (reader.fieldnames or []) if header]

    if not headers:
        raise ValueError("Could not find column headers in this CSV file.")

    columns = _detect_columns(headers)
    if not columns["date"]:
        raise ValueError("Could not find a date column. Expected headers like Date or Transaction Date.")
    if not (columns["amount"] or columns["debit"] or columns["credit"]):
        raise ValueError("Could not find an amount column. Expected Amount, Debit, or Credit.")

    normalized: list[dict[str, str]] = []
    for row in reader:
        if not any(value and str(value).strip() for value in row.values()):
            continue
        normalized.append(normalize_row(row, columns))

    return normalized


def load_transactions(source: Path | str | TextIO, expenses_only: bool = True) -> list[dict]:
    raw_text = _read_csv_text(source)
    rows = parse_csv_rows(raw_text)

    if not rows:
        raise ValueError("No rows found in CSV file.")

    if expenses_only:
        # Keep purchases and refunds; hide card payments / transfers.
        rows = [
            row
            for row in rows
            if row.get("Type", "").strip() in {"Expense", "Return"}
        ]

    if not rows:
        raise ValueError(
            "No expense transactions found. Try turning off 'Expenses only' or check the file format."
        )

    return rows


def split_by_sign(transactions: list[dict]) -> tuple[list[dict], list[dict]]:
    charges: list[dict] = []
    credits: list[dict] = []
    for row in transactions:
        amount = parse_amount(row["Amount"])
        if amount < 0:
            charges.append(row)
        elif amount > 0:
            credits.append(row)
    return charges, credits


def serialize_transaction(row: dict) -> dict:
    amount = parse_amount(row["Amount"])
    return {
        "date": row.get("Date", ""),
        "description": row.get("Description", ""),
        "amount": amount,
        "category": row.get("Category", "").strip() or "Uncategorized",
        "color": category_color(row.get("Category", "")),
        "account": row.get("Account", ""),
        "source": row.get("_source", ""),
        "type": row.get("Type", ""),
        "is_credit": amount > 0,
    }


def group_by_category(transactions: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in transactions:
        category = row.get("Category", "").strip() or "Uncategorized"
        grouped[category].append(row)
    return dict(grouped)


def category_total(transactions: list[dict]) -> float:
    return sum(parse_amount(row["Amount"]) for row in transactions)


def sorted_categories(grouped: dict[str, list[dict]]) -> list[str]:
    return sorted(
        grouped.keys(),
        key=lambda name: abs(category_total(grouped[name])),
        reverse=True,
    )


def build_monthly_summary(transactions: list[dict]) -> dict[str, dict[str, float]]:
    monthly: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for row in transactions:
        month = parse_month(row.get("Date", ""))
        category = row.get("Category", "").strip() or "Uncategorized"
        monthly[month][category] += parse_amount(row["Amount"])
        monthly[month]["_total"] += parse_amount(row["Amount"])

    return {month: dict(categories) for month, categories in sorted(monthly.items(), reverse=True)}


def category_color(name: str) -> str:
    return CATEGORY_COLORS.get(name, "#64748b")


def build_category_entry(name: str, items: list[dict]) -> dict:
    total = category_total(items)
    return {
        "name": name,
        "total": abs(total),
        "signed_total": total,
        "count": len(items),
        "color": category_color(name),
        "used": len(items) > 0,
        "transactions": [
            serialize_transaction(row)
            for row in sorted(items, key=lambda row: row.get("Date", ""), reverse=True)
        ],
    }


def full_category_list(grouped: dict[str, list[dict]]) -> list[dict]:
    seen = set(grouped.keys())
    categories = []

    for name in sorted_categories(grouped):
        categories.append(build_category_entry(name, grouped[name]))
        seen.add(name)

    for name in STANDARD_CATEGORIES:
        if name not in seen:
            categories.append(build_category_entry(name, []))

    return categories


def build_calendar_summaries(transactions: list[dict]) -> list[dict]:
    from calendar import monthrange

    by_date: dict[str, list[dict]] = defaultdict(list)
    for row in transactions:
        date = normalize_date(row.get("Date", ""))
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            continue
        by_date[date].append(row)

    by_month: dict[str, dict[str, list[dict]]] = defaultdict(dict)
    for date, items in by_date.items():
        by_month[date[:7]][date] = items

    calendars: list[dict] = []
    for month in sorted(by_month.keys(), reverse=True):
        year, mon = map(int, month.split("-"))
        days_in_month = monthrange(year, mon)[1]
        first_weekday = (datetime(year, mon, 1).weekday() + 1) % 7

        days: list[dict] = []
        for day_num in range(1, days_in_month + 1):
            date_str = f"{month}-{day_num:02d}"
            items = by_month[month].get(date_str, [])
            charges = [item for item in items if parse_amount(item["Amount"]) < 0]
            credits = [item for item in items if parse_amount(item["Amount"]) > 0]
            spent = round(sum(abs(parse_amount(item["Amount"])) for item in charges), 2)
            credited = round(sum(parse_amount(item["Amount"]) for item in credits), 2)
            net = round(category_total(items), 2)
            days.append(
                {
                    "date": date_str,
                    "day": day_num,
                    "total": spent,
                    "spent": spent,
                    "credits": credited,
                    "net": net,
                    "count": len(items),
                    "transactions": [
                        serialize_transaction(item)
                        for item in sorted(
                            items,
                            key=lambda row: (parse_amount(row["Amount"]) > 0, row.get("Description", "")),
                        )
                    ],
                }
            )

        month_spent = round(sum(day["spent"] for day in days), 2)
        month_credits = round(sum(day["credits"] for day in days), 2)
        month_net = round(sum(day["net"] for day in days), 2)

        calendars.append(
            {
                "month": month,
                "label": datetime(year, mon, 1).strftime("%B %Y"),
                "short_label": datetime(year, mon, 1).strftime("%B").upper(),
                "total": month_spent,
                "spent": month_spent,
                "credits": month_credits,
                "net": month_net,
                "first_weekday": first_weekday,
                "days_in_month": days_in_month,
                "days": days,
            }
        )

    return calendars


def analyze(transactions: list[dict]) -> dict:
    charges, credits = split_by_sign(transactions)
    grouped = group_by_category(charges)
    all_cats = full_category_list(grouped)
    used_categories = [cat for cat in all_cats if cat["used"]]
    unused_categories = [cat for cat in all_cats if not cat["used"]]

    spent_total = round(sum(abs(parse_amount(row["Amount"])) for row in charges), 2)
    credits_total = round(sum(parse_amount(row["Amount"]) for row in credits), 2)
    net_total = round(category_total(transactions), 2)

    monthly_raw = build_monthly_summary(transactions)
    monthly = []
    for month, totals in monthly_raw.items():
        month_net = round(totals.pop("_total", 0.0), 2)
        month_rows = [row for row in transactions if parse_month(row.get("Date", "")) == month]
        month_charges, month_credits = split_by_sign(month_rows)
        month_spent = round(sum(abs(parse_amount(row["Amount"])) for row in month_charges), 2)
        month_credited = round(sum(parse_amount(row["Amount"]) for row in month_credits), 2)

        charge_totals: dict[str, float] = defaultdict(float)
        for row in month_charges:
            category = row.get("Category", "").strip() or "Uncategorized"
            charge_totals[category] += abs(parse_amount(row["Amount"]))

        breakdown = [
            {"category": category, "total": round(amount, 2), "color": category_color(category)}
            for category, amount in sorted(charge_totals.items(), key=lambda item: item[1], reverse=True)
        ]
        monthly.append(
            {
                "month": month,
                "total": month_spent,
                "spent": month_spent,
                "credits": month_credited,
                "net": month_net,
                "categories": breakdown,
            }
        )

    credit_transactions = [
        serialize_transaction(row)
        for row in sorted(credits, key=lambda row: row.get("Date", ""), reverse=True)
    ]

    return {
        "grand_total": spent_total,
        "spent_total": spent_total,
        "credits_total": credits_total,
        "net_total": net_total,
        "transaction_count": len(transactions),
        "charge_count": len(charges),
        "credit_count": len(credits),
        "categories": used_categories,
        "unused_categories": unused_categories,
        "credits": credit_transactions,
        "all_categories": STANDARD_CATEGORIES,
        "monthly": monthly,
        "calendars": build_calendar_summaries(transactions),
    }


def merge_transaction_lists(sources: list[tuple[str, list[dict]]]) -> list[dict]:
    merged: list[dict] = []
    for source_name, rows in sources:
        for row in rows:
            tagged = dict(row)
            tagged["_source"] = source_name
            merged.append(tagged)
    return merged


def build_report(grouped: dict[str, list[dict]], monthly: dict[str, dict[str, float]] | None = None) -> str:
    lines: list[str] = []
    grand_total = 0.0

    if monthly:
        lines.append("MONTHLY SUMMARY")
        lines.append("=" * 72)
        for month, totals in sorted(monthly.items(), reverse=True):
            month_total = totals.get("_total", 0.0)
            lines.append(f"\n{month} — Total: {format_money(month_total)}")
            for category, amount in sorted(
                ((k, v) for k, v in totals.items() if k != "_total"),
                key=lambda item: abs(item[1]),
                reverse=True,
            ):
                lines.append(f"  {category:<28} {format_money(amount):>12}")
        lines.append("")
        lines.append("")

    for category in sorted_categories(grouped):
        items = sorted(grouped[category], key=lambda row: row.get("Date", ""), reverse=True)
        total = category_total(items)
        grand_total += total

        lines.append("=" * 72)
        lines.append(f"{category.upper()} — Total: {format_money(total)} ({len(items)} transactions)")
        lines.append("=" * 72)

        for row in items:
            date = row.get("Date", "")
            description = row.get("Description", "")
            amount = parse_amount(row["Amount"])
            account = row.get("Account", "")
            lines.append(f"  {date}  {format_money(amount):>10}  {description}")
            if account:
                lines.append(f"             {account}")

        lines.append("")

    lines.append("=" * 72)
    lines.append(f"GRAND TOTAL (expenses): {format_money(grand_total)}")
    lines.append("=" * 72)

    return "\n".join(lines)


def export_excel(transactions: list[dict], output: Path | BinaryIO) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    charges, credits = split_by_sign(transactions)
    grouped = group_by_category(charges)
    analysis = analyze(transactions)

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Category Summary"
    summary_sheet.append(["Category", "Transactions", "Spent"])
    for category in analysis["categories"]:
        summary_sheet.append([category["name"], category["count"], category["total"]])
    summary_sheet.append([])
    summary_sheet.append(["Total Spent", analysis["charge_count"], analysis["spent_total"]])
    summary_sheet.append(["Total Credits", analysis["credit_count"], analysis["credits_total"]])
    summary_sheet.append(["Net", analysis["transaction_count"], analysis["net_total"]])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for column in ("A", "B", "C"):
        summary_sheet.column_dimensions[column].width = 24 if column == "A" else 16
    for row in summary_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00;$ (#,##0.00);"-"'

    monthly_sheet = workbook.create_sheet("Monthly Summary")
    monthly_sheet.append(["Month", "Category", "Spent", "Credits", "Net"])
    for entry in analysis["monthly"]:
        for item in entry["categories"]:
            monthly_sheet.append([entry["month"], item["category"], item["total"], "", ""])
        monthly_sheet.append(
            [entry["month"], "MONTH TOTAL", entry["spent"], entry["credits"], entry["net"]]
        )
        monthly_sheet.append([])

    for cell in monthly_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for column, width in (("A", 14), ("B", 28), ("C", 14), ("D", 14), ("E", 14)):
        monthly_sheet.column_dimensions[column].width = width
    for row in monthly_sheet.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00;$ (#,##0.00);"-"'

    if credits:
        credit_sheet = workbook.create_sheet("Credits")
        credit_sheet.append(["Date", "Description", "Category", "Amount", "Account"])
        for cell in credit_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in sorted(credits, key=lambda item: item.get("Date", ""), reverse=True):
            credit_sheet.append(
                [
                    row.get("Date", ""),
                    row.get("Description", ""),
                    row.get("Category", "") or "Uncategorized",
                    parse_amount(row["Amount"]),
                    row.get("Account", ""),
                ]
            )
        for column, width in (("A", 14), ("B", 36), ("C", 24), ("D", 14), ("E", 42)):
            credit_sheet.column_dimensions[column].width = width
        for excel_row in credit_sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in excel_row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "$#,##0.00"

    for category_name in sorted_categories(grouped):
        safe_name = category_name[:31].replace("/", "-")
        sheet = workbook.create_sheet(safe_name)
        sheet.append(["Date", "Description", "Amount", "Account"])
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        items = sorted(grouped[category_name], key=lambda row: row.get("Date", ""), reverse=True)
        for row in items:
            sheet.append(
                [
                    row.get("Date", ""),
                    row.get("Description", ""),
                    parse_amount(row["Amount"]),
                    row.get("Account", ""),
                ]
            )
        sheet.append([])
        sheet.append(["", "CATEGORY TOTAL", category_total(items), ""])

        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 36
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 42
        for excel_row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in excel_row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.00;$ (#,##0.00);"-"'

    if isinstance(output, Path):
        workbook.save(output)
    else:
        workbook.save(output)
